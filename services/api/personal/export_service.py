# CineVault OS — Multi-Format Personal Data Export Service (W8)
# Supports JSON v2.0 (lossless), CSV ZIP (relational), Excel XLSX (openpyxl), and Markdown (.md)

import csv
import io
import json
import zipfile
from datetime import datetime, timezone
from typing import Dict, Any, List, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .mapping import sanitize_formula_injection

def _normalize_export_input(data: Any) -> Dict[str, Any]:
    if hasattr(data, "model_dump"):
        return data.model_dump()
    elif hasattr(data, "dict"):
        return data.dict()
    elif isinstance(data, dict):
        return data
    return {}

# ── JSON Export Builder ─────────────────────────────────────────────────────

def build_json_export(export_data: Any) -> str:
    """Produces canonical lossless JSON v2.0 export document."""
    data = _normalize_export_input(export_data)
    payload = {
        "export_metadata": {
            "schema_version": "2.0.0",
            "exported_at": data.get("exported_at") or datetime.now(timezone.utc).isoformat(),
            "generator": "CineVault OS v2.0 Portability Engine",
            "format": "json",
            "user_id": data.get("user_id", "")
        },
        "user_profile": {
            "user_id": data.get("user_id", ""),
            "streak": data.get("user_profile", {}).get("streak") or data.get("streak", {})
        },
        "library": data.get("library", []),
        "watchlist": data.get("watchlist", []),
        "watch_history": data.get("watch_history", []),
        "ratings": data.get("ratings", []),
        "user_title_states": data.get("user_title_states", []),
        "private_notes": data.get("private_notes", []),
        "reviews": data.get("reviews", []),
        "collections": data.get("custom_lists", [])
    }
    return json.dumps(payload, indent=2, ensure_ascii=False)


def build_csv_zip_export(export_data: Any) -> bytes:
    """
    Produces a ZIP archive containing individual relational CSV tables:
    - library.csv
    - watch_history.csv
    - ratings.csv
    - notes.csv
    - reviews.csv
    - collections.csv
    - manifest.json
    All values are sanitized against spreadsheet formula injection.
    """
    export_data = _normalize_export_input(export_data)
    zip_buffer = io.BytesIO()

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        # 1. Manifest
        manifest = {
            "schema_version": "2.0.0",
            "exported_at": export_data.get("exported_at") or datetime.now(timezone.utc).isoformat(),
            "user_id": export_data.get("user_id", ""),
            "tables": ["library.csv", "watch_history.csv", "ratings.csv", "notes.csv", "reviews.csv", "collections.csv"]
        }
        zf.writestr("manifest.json", json.dumps(manifest, indent=2))

        # 2. library.csv
        lib_buf = io.StringIO()
        lib_writer = csv.writer(lib_buf)
        lib_writer.writerow(["title_id", "display_id", "canonical_title", "production_year", "content_type", "added_at", "status_override"])
        for item in export_data.get("library", []):
            lib_writer.writerow([
                item.get("title_id", ""),
                item.get("display_id", ""),
                sanitize_formula_injection(item.get("canonical_title", "")),
                item.get("production_year") or "",
                item.get("content_type", "movie"),
                item.get("added_at", ""),
                item.get("status_override", "")
            ])
        zf.writestr("library.csv", lib_buf.getvalue())

        # 3. watch_history.csv
        wh_buf = io.StringIO()
        wh_writer = csv.writer(wh_buf)
        wh_writer.writerow([
            "watch_event_id", "title_id", "display_id", "canonical_title", "production_year",
            "content_type", "season_number", "episode_number", "episode_name",
            "watched_at", "device_type", "notes"
        ])
        for we in export_data.get("watch_history", []):
            wh_writer.writerow([
                we.get("watch_event_id", ""),
                we.get("title_id", ""),
                we.get("display_id", ""),
                sanitize_formula_injection(we.get("canonical_title", "")),
                we.get("production_year") or "",
                we.get("content_type", "movie"),
                we.get("season_number") or "",
                we.get("episode_number") or "",
                sanitize_formula_injection(we.get("episode_name", "")),
                we.get("watched_at", ""),
                we.get("device_type", ""),
                sanitize_formula_injection(we.get("notes", ""))
            ])
        zf.writestr("watch_history.csv", wh_buf.getvalue())

        # 4. ratings.csv
        r_buf = io.StringIO()
        r_writer = csv.writer(r_buf)
        r_writer.writerow(["rating_id", "title_id", "display_id", "canonical_title", "production_year", "rating_value", "rated_at"])
        for r in export_data.get("ratings", []):
            r_writer.writerow([
                r.get("rating_id", ""),
                r.get("title_id", ""),
                r.get("display_id", ""),
                sanitize_formula_injection(r.get("canonical_title", "")),
                r.get("production_year") or "",
                r.get("rating_value", ""),
                r.get("rated_at", "")
            ])
        zf.writestr("ratings.csv", r_buf.getvalue())

        # 5. notes.csv
        n_buf = io.StringIO()
        n_writer = csv.writer(n_buf)
        n_writer.writerow(["note_id", "title_id", "display_id", "canonical_title", "production_year", "note_text", "created_at"])
        for n in export_data.get("private_notes", []):
            n_writer.writerow([
                n.get("note_id", ""),
                n.get("title_id", ""),
                n.get("display_id", ""),
                sanitize_formula_injection(n.get("canonical_title", "")),
                n.get("production_year") or "",
                sanitize_formula_injection(n.get("note_text", "")),
                n.get("created_at", "")
            ])
        zf.writestr("notes.csv", n_buf.getvalue())

        # 6. reviews.csv
        rev_buf = io.StringIO()
        rev_writer = csv.writer(rev_buf)
        rev_writer.writerow(["review_id", "title_id", "display_id", "canonical_title", "production_year", "review_title", "review_text", "contains_spoilers", "created_at"])
        for rev in export_data.get("reviews", []):
            rev_writer.writerow([
                rev.get("review_id", ""),
                rev.get("title_id", ""),
                rev.get("display_id", ""),
                sanitize_formula_injection(rev.get("canonical_title", "")),
                rev.get("production_year") or "",
                sanitize_formula_injection(rev.get("review_title", "")),
                sanitize_formula_injection(rev.get("review_text", "")),
                "true" if rev.get("contains_spoilers") else "false",
                rev.get("created_at", "")
            ])
        zf.writestr("reviews.csv", rev_buf.getvalue())

        # 7. collections.csv
        col_buf = io.StringIO()
        col_writer = csv.writer(col_buf)
        col_writer.writerow(["list_id", "collection_title", "description", "is_private", "item_position", "title_id", "display_id", "canonical_title", "production_year", "item_notes"])
        for cl in export_data.get("custom_lists", []):
            list_id = cl.get("list_id", "")
            col_title = cl.get("title", "")
            col_desc = cl.get("description", "")
            is_priv = "true" if cl.get("is_private", True) else "false"
            for item in cl.get("items", []):
                col_writer.writerow([
                    list_id,
                    sanitize_formula_injection(col_title),
                    sanitize_formula_injection(col_desc),
                    is_priv,
                    item.get("position", 0),
                    item.get("title_id", ""),
                    item.get("display_id", ""),
                    sanitize_formula_injection(item.get("canonical_title", "")),
                    item.get("production_year") or "",
                    sanitize_formula_injection(item.get("notes", ""))
                ])
        zf.writestr("collections.csv", col_buf.getvalue())

    zip_buffer.seek(0)
    return zip_buffer.getvalue()


def build_excel_export(export_data: Any) -> bytes:
    """Produces a multi-sheet Excel .xlsx workbook using openpyxl."""
    export_data = _normalize_export_input(export_data)
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4C1D95", end_color="4C1D95", fill_type="solid") # Violet-900
    sub_header_fill = PatternFill(start_color="5B21B6", end_color="5B21B6", fill_type="solid") # Violet-800
    regular_font = Font(name="Segoe UI", size=10)
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB")
    )

    def style_table(ws, headers: List[str], rows: List[List[Any]]):
        ws.append(headers)
        header_row = ws[1]
        for cell in header_row:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")

        for row_idx, r in enumerate(rows, start=2):
            sanitized_row = [sanitize_formula_injection(val) if isinstance(val, str) else val for val in r]
            ws.append(sanitized_row)
            for col_idx in range(1, len(headers) + 1):
                c = ws.cell(row=row_idx, column=col_idx)
                c.font = regular_font
                c.border = thin_border

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # 1. Overview Sheet
    ws_over = wb.create_sheet(title="Overview")
    ws_over.views.sheetView[0].showGridLines = True
    ws_over.append(["CineVault OS — Personal Vault Archive", ""])
    ws_over["A1"].font = Font(name="Segoe UI", size=14, bold=True, color="4C1D95")
    ws_over.append(["Export Schema Version", "2.0.0"])
    ws_over.append(["Exported At (UTC)", export_data.get("exported_at", "")])
    ws_over.append(["User ID", export_data.get("user_id", "")])
    ws_over.append(["", ""])
    ws_over.append(["Domain Summary", "Record Count"])
    ws_over["A6"].font = header_font
    ws_over["A6"].fill = sub_header_fill
    ws_over["B6"].font = header_font
    ws_over["B6"].fill = sub_header_fill
    ws_over.append(["Library Entries", len(export_data.get("library", []))])
    ws_over.append(["Watch Events Logged", len(export_data.get("watch_history", []))])
    ws_over.append(["Ratings Assigned", len(export_data.get("ratings", []))])
    ws_over.append(["Private Notes", len(export_data.get("private_notes", []))])
    ws_over.append(["Reviews", len(export_data.get("reviews", []))])
    ws_over.append(["Custom Collections", len(export_data.get("custom_lists", []))])

    streak = export_data.get("streak", {})
    if streak:
        ws_over.append(["Current Watch Streak (Days)", streak.get("current_streak", 0)])
        ws_over.append(["Longest Watch Streak (Days)", streak.get("longest_streak", 0)])

    ws_over.column_dimensions["A"].width = 30
    ws_over.column_dimensions["B"].width = 40

    # 2. Library & Watchlist Sheet
    ws_lib = wb.create_sheet(title="Library & Watchlist")
    lib_headers = ["Title", "Year", "Type", "Display ID", "Status Override", "Added At"]
    lib_rows = [
        [
            item.get("canonical_title", ""),
            item.get("production_year") or "",
            item.get("content_type", "movie"),
            item.get("display_id", ""),
            item.get("status_override", ""),
            item.get("added_at", "")
        ]
        for item in export_data.get("library", [])
    ]
    style_table(ws_lib, lib_headers, lib_rows)

    # 3. Watch Events Sheet
    ws_we = wb.create_sheet(title="Watch Events")
    we_headers = ["Title", "Year", "Type", "Season", "Episode", "Episode Title", "Watched Date", "Device", "Notes"]
    we_rows = [
        [
            we.get("canonical_title", ""),
            we.get("production_year") or "",
            we.get("content_type", "movie"),
            we.get("season_number") or "",
            we.get("episode_number") or "",
            we.get("episode_name", ""),
            we.get("watched_at", ""),
            we.get("device_type", ""),
            we.get("notes", "")
        ]
        for we in export_data.get("watch_history", [])
    ]
    style_table(ws_we, we_headers, we_rows)

    # 4. Ratings Sheet
    ws_r = wb.create_sheet(title="Ratings")
    r_headers = ["Title", "Year", "Rating (1-10)", "Rated Date"]
    r_rows = [
        [
            r.get("canonical_title", ""),
            r.get("production_year") or "",
            r.get("rating_value", ""),
            r.get("rated_at", "")
        ]
        for r in export_data.get("ratings", [])
    ]
    style_table(ws_r, r_headers, r_rows)

    # 5. Notes & Reviews Sheet
    ws_n = wb.create_sheet(title="Notes & Reviews")
    nr_headers = ["Type", "Title", "Year", "Header / Summary", "Text Content", "Spoilers", "Created Date"]
    nr_rows = []
    for n in export_data.get("private_notes", []):
        nr_rows.append([
            "Private Note",
            n.get("canonical_title", ""),
            n.get("production_year") or "",
            "-",
            n.get("note_text", ""),
            "No",
            n.get("created_at", "")
        ])
    for rev in export_data.get("reviews", []):
        nr_rows.append([
            "Review",
            rev.get("canonical_title", ""),
            rev.get("production_year") or "",
            rev.get("review_title", ""),
            rev.get("review_text", ""),
            "Yes" if rev.get("contains_spoilers") else "No",
            rev.get("created_at", "")
        ])
    style_table(ws_n, nr_headers, nr_rows)

    # 6. Collections Sheet
    ws_col = wb.create_sheet(title="Collections")
    col_headers = ["Collection Name", "Description", "Privacy", "Position", "Title", "Year", "Item Notes"]
    col_rows = []
    for cl in export_data.get("custom_lists", []):
        c_name = cl.get("title", "")
        c_desc = cl.get("description", "")
        c_priv = "Private" if cl.get("is_private", True) else "Public"
        for it in cl.get("items", []):
            col_rows.append([
                c_name,
                c_desc,
                c_priv,
                it.get("position", 0),
                it.get("canonical_title", ""),
                it.get("production_year") or "",
                it.get("notes", "")
            ])
    style_table(ws_col, col_headers, col_rows)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


# ── Markdown Export Builder ─────────────────────────────────────────────────

def build_markdown_export(export_data: Any) -> str:
    """Produces human-readable Markdown personal archive document."""
    export_data = _normalize_export_input(export_data)
    exported_at = export_data.get("exported_at") or datetime.now(timezone.utc).isoformat()
    user_id = export_data.get("user_id", "")
    lines: List[str] = [
        f"# CineVault Personal Media Archive",
        f"",
        f"- **Export Date (UTC):** `{exported_at}`",
        f"- **Schema Version:** `2.0.0`",
        f"- **User Identifier:** `{user_id}`",
        f"",
        f"---",
        f"",
        f"## 📊 Archive Summary",
        f"",
        f"| Domain | Total Records |",
        f"| :--- | :--- |",
        f"| **Library Titles** | {len(export_data.get('library', []))} |",
        f"| **Watch History Events** | {len(export_data.get('watch_history', []))} |",
        f"| **Ratings Assigned** | {len(export_data.get('ratings', []))} |",
        f"| **Private Notes** | {len(export_data.get('private_notes', []))} |",
        f"| **Reviews Written** | {len(export_data.get('reviews', []))} |",
        f"| **Custom Collections** | {len(export_data.get('custom_lists', []))} |",
        f"",
    ]

    # 1. Library & Watchlist
    lines.extend([
        f"## 📚 Library & Watchlist",
        f"",
    ])
    library = export_data.get("library", [])
    if library:
        lines.append("| Title | Year | Type | Display ID | Status | Added At |")
        lines.append("| :--- | :--- | :--- | :--- | :--- | :--- |")
        for item in library:
            title = item.get("canonical_title", "Unknown Title")
            year = item.get("production_year", "—")
            ctype = item.get("content_type", "movie").upper()
            did = item.get("display_id", "—")
            status = item.get("status_override", "LIBRARY")
            added = item.get("added_at", "—")[:10] if item.get("added_at") else "—"
            lines.append(f"| **{title}** | {year} | {ctype} | `{did}` | {status} | {added} |")
    else:
        lines.append("*No items in library.*")
    lines.append("")

    # 2. Watch History
    lines.extend([
        f"## ⏱️ Watch History",
        f"",
    ])
    watch_history = export_data.get("watch_history", [])
    if watch_history:
        lines.append("| Title | Episode | Watched Date | Device | Notes |")
        lines.append("| :--- | :--- | :--- | :--- | :--- |")
        for we in watch_history:
            title = we.get("canonical_title", "Unknown Title")
            year = f" ({we.get('production_year')})" if we.get("production_year") else ""
            ep_str = "—"
            if we.get("season_number") is not None and we.get("episode_number") is not None:
                ep_str = f"S{we.get('season_number'):02d}:E{we.get('episode_number'):02d}"
                if we.get("episode_name"):
                    ep_str += f" - {we.get('episode_name')}"
            w_date = we.get("watched_at", "—")[:10] if we.get("watched_at") else "—"
            device = we.get("device_type", "—") or "—"
            notes = we.get("notes", "—") or "—"
            lines.append(f"| **{title}{year}** | {ep_str} | {w_date} | {device} | {notes} |")
    else:
        lines.append("*No watch events logged.*")
    lines.append("")

    # 3. Ratings
    lines.extend([
        f"## ⭐ Ratings",
        f"",
    ])
    ratings = export_data.get("ratings", [])
    if ratings:
        lines.append("| Title | Year | Score | Rated Date |")
        lines.append("| :--- | :--- | :--- | :--- |")
        for r in ratings:
            title = r.get("canonical_title", "Unknown Title")
            year = r.get("production_year", "—")
            score = f"★ {r.get('rating_value')}/10"
            r_date = r.get("rated_at", "—")[:10] if r.get("rated_at") else "—"
            lines.append(f"| **{title}** | {year} | {score} | {r_date} |")
    else:
        lines.append("*No ratings assigned.*")
    lines.append("")

    # 4. Notes
    lines.extend([
        f"## 📝 Private Notes",
        f"",
    ])
    notes = export_data.get("private_notes", [])
    if notes:
        for n in notes:
            title = n.get("canonical_title", "Unknown Title")
            year = f" ({n.get('production_year')})" if n.get("production_year") else ""
            date_str = n.get("created_at", "")[:10]
            lines.append(f"### {title}{year} — *{date_str}*")
            lines.append(f"")
            lines.append(f"> {n.get('note_text', '')}")
            lines.append(f"")
    else:
        lines.append("*No private notes recorded.*")
        lines.append("")

    # 5. Reviews
    lines.extend([
        f"## ✍️ Reviews",
        f"",
    ])
    reviews = export_data.get("reviews", [])
    if reviews:
        for rev in reviews:
            title = rev.get("canonical_title", "Unknown Title")
            year = f" ({rev.get('production_year')})" if rev.get("production_year") else ""
            rev_title = rev.get("review_title", "Review")
            spoiler = " ⚠️ [Contains Spoilers]" if rev.get("contains_spoilers") else ""
            lines.append(f"### {title}{year}: \"{rev_title}\"{spoiler}")
            lines.append(f"")
            lines.append(rev.get("review_text", ""))
            lines.append(f"")
    else:
        lines.append("*No reviews recorded.*")
        lines.append("")

    # 6. Collections
    lines.extend([
        f"## 📁 Custom Collections",
        f"",
    ])
    collections = export_data.get("custom_lists", [])
    if collections:
        for cl in collections:
            c_title = cl.get("title", "Untitled Collection")
            c_desc = f" — *{cl.get('description')}*" if cl.get("description") else ""
            lines.append(f"### 📂 {c_title}{c_desc}")
            lines.append(f"")
            items = cl.get("items", [])
            if items:
                for it in items:
                    pos = it.get("position", 0) + 1
                    t_name = it.get("canonical_title", "Unknown")
                    y = f" ({it.get('production_year')})" if it.get("production_year") else ""
                    it_notes = f" — {it.get('notes')}" if it.get("notes") else ""
                    lines.append(f"{pos}. **{t_name}{y}**{it_notes}")
            else:
                lines.append("*Collection is empty.*")
            lines.append(f"")
    else:
        lines.append("*No custom collections created.*")
        lines.append("")

    lines.append("---")
    lines.append("*Export generated by CineVault OS Data Portability Engine.*")

    return "\n".join(lines)
