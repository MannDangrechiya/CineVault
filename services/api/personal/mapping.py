# CineVault OS — Personal Data Import Mapping & Parsing Layer
# Enforces safe header mapping, formula injection neutralization, and multi-format parsing

import csv
import io
import json
import re
from datetime import datetime, timezone
from typing import List, Dict, Any, Optional, Tuple
import openpyxl

# ── Formula Injection Defense ───────────────────────────────────────────────
DANGEROUS_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")

def sanitize_formula_injection(value: Any) -> Any:
    """Neutralizes spreadsheet formula injection by prefixing risky strings with a single quote."""
    if isinstance(value, str):
        # If string begins with a formula trigger character
        if value.startswith(DANGEROUS_FORMULA_PREFIXES):
            return f"'{value}"
    return value

def strip_formula_prefix(value: Any) -> Any:
    """Strips leading single quote used for formula escaping when reading values."""
    if isinstance(value, str) and value.startswith("'") and len(value) > 1:
        if value[1] in DANGEROUS_FORMULA_PREFIXES:
            return value[1:]
    return value


# ── Field Alias Mapping ─────────────────────────────────────────────────────
# Maps common external column names across Letterboxd, Trakt, IMDb, and spreadsheets to CineVault fields.

FIELD_ALIASES: Dict[str, List[str]] = {
    "title": [
        "title", "canonical_title", "film", "movie", "movie name", "movie_title",
        "film title", "film_title", "series title", "show name", "name", "original title"
    ],
    "year": [
        "year", "production_year", "release year", "release_year", "prod_year", "released", "year of release"
    ],
    "rating": [
        "rating", "rating_value", "rating10", "score", "user_rating", "your rating",
        "stars", "letterboxd rating", "my rating"
    ],
    "watched_at": [
        "watched_at", "watched date", "date", "watched", "viewed_at", "log_date",
        "logged date", "watched on", "view date", "viewed date"
    ],
    "notes": [
        "notes", "note", "comment", "comments", "review", "review_text", "tags",
        "personal notes", "user note"
    ],
    "status": [
        "status", "manual_status_override", "state", "list", "shelf", "watch_status"
    ],
    "favorite": [
        "favorite", "is_favorite", "fav", "liked", "like"
    ],
    "imdb_id": [
        "imdb_id", "imdb id", "const", "imdb_title_id", "tconst"
    ],
    "tmdb_id": [
        "tmdb_id", "tmdb id", "themoviedb_id"
    ],
    "display_id": [
        "display_id", "display id", "cinevault_id", "id"
    ],
    "title_id": [
        "title_id", "canonical_id", "uuid"
    ],
    "season_number": [
        "season", "season_number", "season_num", "s"
    ],
    "episode_number": [
        "episode", "episode_number", "episode_num", "ep", "e"
    ]
}


def normalize_header(header: str) -> str:
    """Normalizes header string for alias lookup."""
    clean = header.strip().lower().replace("_", " ").replace("-", " ")
    clean = re.sub(r"\s+", " ", clean)
    return clean


def resolve_field_name(raw_header: str) -> Optional[str]:
    """Resolves an external column header to a canonical CineVault field name."""
    norm = normalize_header(raw_header)
    for canonical_field, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            if norm == alias or norm == alias.replace(" ", ""):
                return canonical_field
    return None


# ── Value Normalizers ───────────────────────────────────────────────────────

def parse_year(val: Any) -> Optional[int]:
    """Parses year integer safely."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        try:
            y = int(val)
            if 1850 <= y <= 2100:
                return y
        except (ValueError, OverflowError):
            return None
    s = str(val).strip()
    match = re.search(r"\b(18\d\d|19\d\d|20\d\d)\b", s)
    if match:
        return int(match.group(1))
    return None


def parse_rating(val: Any) -> Optional[int]:
    """
    Normalizes rating to 1-10 integer scale.
    Handles:
    - 5-star scale (e.g. 5 or 4.5 -> converted or scaled to 10)
    - 10-point scale (e.g. 8/10 -> 8)
    - Letterboxd star strings (★★★★★)
    """
    if val is None:
        return None

    if isinstance(val, (int, float)):
        if val <= 5.0:
            scaled = round(val * 2)
            return max(1, min(10, int(scaled))) if scaled > 0 else None
        else:
            scaled = round(val)
            return max(1, min(10, int(scaled)))

    s = str(val).strip()
    stars = s.count("★")
    if stars > 0:
        return max(1, min(10, stars * 2))

    ratio_match = re.match(r"^(\d+(?:\.\d+)?)\s*/\s*(5|10)$", s)
    if ratio_match:
        score = float(ratio_match.group(1))
        max_score = float(ratio_match.group(2))
        if max_score == 5:
            return max(1, min(10, round(score * 2)))
        return max(1, min(10, round(score)))

    try:
        num = float(s)
        if num <= 5.0:
            return max(1, min(10, round(num * 2)))
        return max(1, min(10, round(num)))
    except ValueError:
        return None


def parse_timestamp(val: Any) -> Optional[str]:
    """Parses date/datetime to ISO-8601 UTC string."""
    if val is None:
        return None
    if isinstance(val, datetime):
        if val.tzinfo is None:
            val = val.replace(tzinfo=timezone.utc)
        return val.isoformat()

    s = str(val).strip()
    if not s:
        return None

    try:
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.isoformat()
    except Exception:
        pass

    date_match = re.match(r"^(\d{4})[/-](\d{1,2})[/-](\d{1,2})$", s)
    if date_match:
        y, m, d = int(date_match.group(1)), int(date_match.group(2)), int(date_match.group(3))
        dt = datetime(y, m, d, tzinfo=timezone.utc)
        return dt.isoformat()

    return None


def parse_status(val: Any) -> str:
    """Normalizes status override strings."""
    if not val:
        return "COMPLETED"
    s = str(val).strip().upper()
    if "WATCHLIST" in s or "PLAN" in s or "WANT" in s:
        return "PLAN_TO_WATCH"
    if "WATCHING" in s or "CURRENT" in s or "PROGRESS" in s:
        return "WATCHING"
    if "DROP" in s or "ABANDON" in s:
        return "DROPPED"
    return "COMPLETED"


# ── Multi-Format Parsers ────────────────────────────────────────────────────

def parse_csv_content(csv_text: str) -> List[Dict[str, Any]]:
    """Parses CSV text into standardized raw dict rows with resolved field mappings."""
    if not csv_text.strip():
        return []

    sample = csv_text[:2048]
    delimiter = ","
    if "\t" in sample and sample.count("\t") > sample.count(","):
        delimiter = "\t"
    elif ";" in sample and sample.count(";") > sample.count(","):
        delimiter = ";"

    reader = csv.reader(io.StringIO(csv_text), delimiter=delimiter)
    try:
        headers = next(reader)
    except StopIteration:
        return []

    field_map: Dict[int, str] = {}
    for idx, h in enumerate(headers):
        canon_field = resolve_field_name(h)
        if canon_field:
            field_map[idx] = canon_field

    items: List[Dict[str, Any]] = []
    for row in reader:
        if not any(cell.strip() for cell in row):
            continue
        row_dict: Dict[str, Any] = {}
        for idx, cell in enumerate(row):
            if idx in field_map:
                field = field_map[idx]
                row_dict[field] = strip_formula_prefix(cell.strip())
        if row_dict:
            items.append(row_dict)

    return items


def parse_xlsx_content(file_bytes: bytes) -> List[Dict[str, Any]]:
    """Parses an Excel .xlsx workbook into raw dict rows with resolved field mappings."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    items: List[Dict[str, Any]] = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            continue

        header_row_idx = 0
        headers = [str(cell) if cell is not None else "" for cell in rows[0]]

        field_map: Dict[int, str] = {}
        for idx, h in enumerate(headers):
            canon_field = resolve_field_name(h)
            if canon_field:
                field_map[idx] = canon_field

        if not field_map and len(rows) > 2:
            header_row_idx = 1
            headers = [str(cell) if cell is not None else "" for cell in rows[1]]
            for idx, h in enumerate(headers):
                canon_field = resolve_field_name(h)
                if canon_field:
                    field_map[idx] = canon_field

        if not field_map:
            continue

        for row in rows[header_row_idx + 1:]:
            if not row or not any(cell is not None for cell in row):
                continue
            row_dict: Dict[str, Any] = {}
            for idx, cell in enumerate(row):
                if idx in field_map and cell is not None:
                    field = field_map[idx]
                    row_dict[field] = strip_formula_prefix(cell)
            if row_dict:
                items.append(row_dict)

    return items


def parse_json_content(json_text: str) -> List[Dict[str, Any]]:
    """Parses JSON text into standardized raw dict rows (handles array of items or full export format)."""
    parsed = json.loads(json_text)
    if isinstance(parsed, list):
        return parsed
    elif isinstance(parsed, dict):
        combined: List[Dict[str, Any]] = []
        if "library" in parsed and isinstance(parsed["library"], list):
            for item in parsed["library"]:
                item["_source_section"] = "library"
                combined.append(item)
        if "watchlist" in parsed and isinstance(parsed["watchlist"], list):
            for item in parsed["watchlist"]:
                item["manual_status_override"] = "PLAN_TO_WATCH"
                item["_source_section"] = "watchlist"
                combined.append(item)
        if "watch_history" in parsed and isinstance(parsed["watch_history"], list):
            for item in parsed["watch_history"]:
                item["_source_section"] = "watch_history"
                combined.append(item)
        if "ratings" in parsed and isinstance(parsed["ratings"], list):
            for item in parsed["ratings"]:
                item["_source_section"] = "ratings"
                combined.append(item)
        if "private_notes" in parsed and isinstance(parsed["private_notes"], list):
            for item in parsed["private_notes"]:
                item["_source_section"] = "notes"
                combined.append(item)
        if "reviews" in parsed and isinstance(parsed["reviews"], list):
            for item in parsed["reviews"]:
                item["_source_section"] = "reviews"
                combined.append(item)
        if "items" in parsed and isinstance(parsed["items"], list):
            combined.extend(parsed["items"])

        if combined:
            return combined
        return [parsed]
    return []


def parse_unstructured_text_content(text: str) -> List[Dict[str, Any]]:
    """
    Parses unstructured text lists (e.g. Samsung Notes, Apple Notes, Markdown checklists).
    Example lines:
    - 1. Dune: Part Two (2024) - Watched, 5/5 ★★★★★
    - Blade Runner 2049 [2017] ★★★★★ - Roger Deakins masterpiece
    - Oppenheimer (2023) - 9/10 in IMAX 70mm
    """
    lines = text.splitlines()
    items: List[Dict[str, Any]] = []

    for line in lines:
        clean = line.strip()
        clean = re.sub(r"^(\d+[\.\)]|\-|\*|•)\s*", "", clean).strip()
        if not clean:
            continue

        year = None
        rating = None
        notes = None

        stars = clean.count("★")
        if stars > 0:
            rating = max(1, min(10, stars * 2))
            clean = clean.replace("★", "").strip()

        score_match = re.search(r"(\d+(?:\.\d+)?)\s*/\s*(5|10)", clean)
        if score_match:
            val = float(score_match.group(1))
            max_val = float(score_match.group(2))
            rating = round(val * 2) if max_val == 5 else round(val)
            clean = clean[:score_match.start()] + clean[score_match.end():]
            clean = clean.strip()

        year_match = re.search(r"[\(\[]\s*(18\d\d|19\d\d|20\d\d)\s*[\)\]]", clean)
        if year_match:
            year = int(year_match.group(1))
            clean = clean[:year_match.start()] + clean[year_match.end():]
            clean = clean.strip()

        if " - " in clean:
            parts = clean.split(" - ", 1)
            clean = parts[0].strip()
            notes = parts[1].strip()

        clean = re.sub(r"[,:;\-]+$", "", clean).strip()

        if clean:
            items.append({
                "title": clean,
                "year": year,
                "rating": rating,
                "notes": notes,
                "status": "COMPLETED",
                "watched_at": datetime.now(timezone.utc).isoformat()
            })

    return items


def convert_raw_dict_to_import_payload(row: Dict[str, Any]) -> Dict[str, Any]:
    """Converts normalized raw dictionary fields into CineVault ImportItemPayload format."""
    canonical_title = row.get("title") or row.get("canonical_title") or row.get("name") or row.get("film")
    if canonical_title:
        canonical_title = str(canonical_title).strip()

    year = parse_year(row.get("year") or row.get("production_year"))
    rating = parse_rating(row.get("rating") or row.get("rating_value") or row.get("score"))
    watched_at = parse_timestamp(row.get("watched_at") or row.get("date") or row.get("watched_date"))
    status_override = parse_status(row.get("status") or row.get("manual_status_override"))
    is_fav = bool(row.get("favorite") or row.get("is_favorite"))
    notes = row.get("notes") or row.get("note") or row.get("comment")
    if notes:
        notes = str(notes).strip()

    title_id = row.get("title_id") or row.get("canonical_id")
    if title_id:
        title_id = str(title_id).strip()

    imdb_id = row.get("imdb_id") or row.get("const")
    if imdb_id:
        imdb_id = str(imdb_id).strip()

    display_id = row.get("display_id")
    if display_id:
        display_id = str(display_id).strip()

    season_num = parse_year(row.get("season_number"))
    episode_num = parse_year(row.get("episode_number"))

    return {
        "canonical_title": canonical_title,
        "production_year": year,
        "title_id": title_id,
        "watched_at": watched_at or (datetime.now(timezone.utc).isoformat() if status_override == "COMPLETED" else None),
        "rating_value": rating,
        "is_favorite": is_fav,
        "manual_status_override": status_override,
        "notes": notes,
        "imdb_id": imdb_id,
        "display_id": display_id,
        "season_number": season_num,
        "episode_number": episode_num,
    }
